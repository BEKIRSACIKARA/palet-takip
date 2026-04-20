from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import hashlib
import jwt
import datetime
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from functools import wraps
import psycopg2
import psycopg2.extras
import urllib.parse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import zipfile
import csv
import smtplib
import json
import threading
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

app = Flask(__name__, static_folder='static')
app.config['SECRET_KEY'] = 'palet-takip-gizli-anahtar-2026'
CORS(app)

PALET_TIPLERI = [("P001", "Euro Palet"), ("P002", "Sanayi Paleti"), ("P003", "Plastik Palet")]
SAHIP_TIP_DEPO = "DEPO"
SAHIP_TIP_DAGITICI = "DAGITICI"
SAHIP_TIP_MUSTERI = "MUSTERI"

HAREKET_DEPO_DAGITICI = "DEPO_DAGITICI"
HAREKET_DAGITICI_MUSTERI = "DAGITICI_MUSTERI"
HAREKET_MUSTERI_DAGITICI = "MUSTERI_DAGITICI"
HAREKET_DAGITICI_DEPO = "DAGITICI_DEPO"
HAREKET_DEPO_STOK = "DEPO_STOK_HAREKET"

# --- YARDIMCI FONKSİYON: TÜRKİYE SAATİNİ GETİR ---
def get_now():
    return datetime.now() + timedelta(hours=3)


def hash_sifre(sifre):
    return hashlib.sha256(sifre.encode()).hexdigest()


_DB_TYPE = 'postgres'

def get_db_connection():
    global _DB_TYPE
    database_url = os.environ.get('DATABASE_URL')
    if database_url:
        urllib.parse.uses_netloc.append('postgres')
        url = urllib.parse.urlparse(database_url)
        _DB_TYPE = 'postgres'
        return psycopg2.connect(
            database=url.path[1:], user=url.username,
            password=url.password, host=url.hostname, port=url.port
        )
    else:
        import sqlite3
        _DB_TYPE = 'sqlite'
        conn = sqlite3.connect('palet_takip.db')
        conn.row_factory = sqlite3.Row
        return conn


def db_execute(cursor, conn, sql, params=()):
    """PostgreSQL %s ve SQLite ? uyumlu execute"""
    if _DB_TYPE == 'sqlite':
        sql = sql.replace('%s', '?')
    cursor.execute(sql, params)
    return cursor


def veritabani_olustur():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS kullanicilar (id SERIAL PRIMARY KEY, kullanici_adi TEXT UNIQUE NOT NULL, sifre TEXT NOT NULL, tip TEXT NOT NULL, ad_soyad TEXT NOT NULL, kisitlamalar TEXT DEFAULT \'\', aktif INTEGER DEFAULT 1)''')
    try:
        cursor.execute("ALTER TABLE kullanicilar ADD COLUMN kisitlamalar TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        conn.rollback()
    try:
        cursor.execute("ALTER TABLE kullanicilar ADD COLUMN aktif INTEGER DEFAULT 1")
        cursor.execute("UPDATE kullanicilar SET aktif = 1 WHERE aktif IS NULL")
        conn.commit()
    except Exception:
        conn.rollback()
    try:
        cursor.execute("ALTER TABLE kullanicilar ADD COLUMN kisitlamalar TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        conn.rollback()
    cursor.execute('''CREATE TABLE IF NOT EXISTS musteriler (id SERIAL PRIMARY KEY, musteri_kodu TEXT UNIQUE NOT NULL, musteri_adi TEXT NOT NULL, tabela_adi TEXT NOT NULL)''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS palet_tipleri (id SERIAL PRIMARY KEY, stok_kodu TEXT UNIQUE NOT NULL, palet_adi TEXT NOT NULL)''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS stoklar (id SERIAL PRIMARY KEY, stok_sahibi_tip TEXT NOT NULL, stok_sahibi_id INTEGER NOT NULL, palet_tipi_id INTEGER NOT NULL, miktar INTEGER DEFAULT 0, FOREIGN KEY (palet_tipi_id) REFERENCES palet_tipleri(id), UNIQUE(stok_sahibi_tip, stok_sahibi_id, palet_tipi_id))''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS hareketler (id SERIAL PRIMARY KEY, tarih TEXT NOT NULL, yapan_kullanici_id INTEGER NOT NULL, hareket_tipi TEXT NOT NULL, gonderen_tip TEXT NOT NULL, gonderen_id INTEGER NOT NULL, alan_tip TEXT NOT NULL, alan_id INTEGER NOT NULL, palet_tipi_id INTEGER NOT NULL, miktar INTEGER NOT NULL, aciklama TEXT, makbuz_no TEXT, FOREIGN KEY (yapan_kullanici_id) REFERENCES kullanicilar(id), FOREIGN KEY (palet_tipi_id) REFERENCES palet_tipleri(id))''')
    try:
        cursor.execute("ALTER TABLE hareketler ADD COLUMN makbuz_no TEXT")
        conn.commit()
    except Exception:
        conn.rollback()
    cursor.execute('''CREATE TABLE IF NOT EXISTS makbuzlar (id SERIAL PRIMARY KEY, makbuz_no TEXT UNIQUE NOT NULL, tarih TEXT NOT NULL, islem_tipi TEXT NOT NULL, gonderen_tip TEXT NOT NULL, gonderen_id INTEGER NOT NULL, gonderen_adi TEXT NOT NULL, alan_tip TEXT NOT NULL, alan_id INTEGER NOT NULL, alan_adi TEXT NOT NULL, toplam_miktar INTEGER NOT NULL, aciklama TEXT, yapan_kullanici_id INTEGER NOT NULL, yapan_kullanici_adi TEXT NOT NULL, FOREIGN KEY (yapan_kullanici_id) REFERENCES kullanicilar(id))''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS makbuz_detaylari (id SERIAL PRIMARY KEY, makbuz_id INTEGER NOT NULL, palet_tipi_id INTEGER NOT NULL, stok_kodu TEXT NOT NULL, palet_adi TEXT NOT NULL, miktar INTEGER NOT NULL, FOREIGN KEY (makbuz_id) REFERENCES makbuzlar(id), FOREIGN KEY (palet_tipi_id) REFERENCES palet_tipleri(id))''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS ayarlar (key TEXT PRIMARY KEY, value TEXT NOT NULL)''')
    conn.commit()
    for stok_kodu, palet_adi in PALET_TIPLERI:
        cursor.execute("INSERT INTO palet_tipleri (stok_kodu, palet_adi) SELECT %s, %s WHERE NOT EXISTS (SELECT 1 FROM palet_tipleri WHERE stok_kodu = %s)", (stok_kodu, palet_adi, stok_kodu))
    cursor.execute("INSERT INTO kullanicilar (kullanici_adi, sifre, tip, ad_soyad) SELECT %s, %s, %s, %s WHERE NOT EXISTS (SELECT 1 FROM kullanicilar WHERE kullanici_adi = %s)", ('depocu', hash_sifre('12345'), 'DEPOCU', 'Ana Depocu', 'depocu'))
    cursor.execute("UPDATE kullanicilar SET sifre = %s WHERE kullanici_adi = 'depocu'", (hash_sifre('12345'),))

    conn.commit()
    cursor.execute('SELECT id FROM palet_tipleri')
    for palet in cursor.fetchall():
        cursor.execute("INSERT INTO stoklar (stok_sahibi_tip, stok_sahibi_id, palet_tipi_id, miktar) SELECT %s, %s, %s, 0 WHERE NOT EXISTS (SELECT 1 FROM stoklar WHERE stok_sahibi_tip = %s AND stok_sahibi_id = %s AND palet_tipi_id = %s)", ('DEPO', 0, palet[0], 'DEPO', 0, palet[0]))
    conn.commit()
    cursor.close()
    conn.close()


def stok_miktari_getir(stok_sahibi_tip, stok_sahibi_id, palet_tipi_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT miktar FROM stoklar WHERE stok_sahibi_tip = %s AND stok_sahibi_id = %s AND palet_tipi_id = %s", (stok_sahibi_tip, stok_sahibi_id, palet_tipi_id))
    sonuc = cursor.fetchone()
    cursor.close()
    conn.close()
    return sonuc[0] if sonuc else 0


def stok_guncelle(stok_sahibi_tip, stok_sahibi_id, palet_tipi_id, degisim):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, miktar FROM stoklar WHERE stok_sahibi_tip = %s AND stok_sahibi_id = %s AND palet_tipi_id = %s", (stok_sahibi_tip, stok_sahibi_id, palet_tipi_id))
    mevcut = cursor.fetchone()
    if mevcut:
        yeni_miktar = mevcut[1] + degisim
        if yeni_miktar < 0:
            cursor.close()
            conn.close()
            return False, "Stok yetersiz!"
        cursor.execute("UPDATE stoklar SET miktar = %s WHERE id = %s", (yeni_miktar, mevcut[0]))
    else:
        if degisim < 0:
            cursor.close()
            conn.close()
            return False, "Stok kaydı bulunamadı!"
        cursor.execute("INSERT INTO stoklar (stok_sahibi_tip, stok_sahibi_id, palet_tipi_id, miktar) VALUES (%s, %s, %s, %s)", (stok_sahibi_tip, stok_sahibi_id, palet_tipi_id, degisim))
    conn.commit()
    cursor.close()
    conn.close()
    return True, ""


def hareket_kaydet(yapan_kullanici_id, hareket_tipi, gonderen_tip, gonderen_id, alan_tip, alan_id, palet_tipi_id, miktar, aciklama="", makbuz_no=None):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO hareketler (tarih, yapan_kullanici_id, hareket_tipi, gonderen_tip, gonderen_id, alan_tip, alan_id, palet_tipi_id, miktar, aciklama, makbuz_no) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                   (get_now().strftime("%Y-%m-%d %H:%M:%S"), yapan_kullanici_id, hareket_tipi, gonderen_tip, gonderen_id, alan_tip, alan_id, palet_tipi_id, miktar, aciklama, makbuz_no))
    conn.commit()
    cursor.close()
    conn.close()


def makbuz_no_olustur():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT MAX(id) FROM makbuzlar")
        son_id = cursor.fetchone()[0]
        yeni_no = (son_id + 1) if son_id else 1
    except:
        yeni_no = 1
    finally:
        cursor.close()
        conn.close()
    return f"PLT{str(yeni_no).zfill(5)}"


def makbuz_kaydet(transfer_data):
    conn = get_db_connection()
    cursor = conn.cursor()
    makbuz_no = makbuz_no_olustur()
    tarih = get_now().strftime("%d.%m.%Y %H:%M:%S")
    cursor.execute("INSERT INTO makbuzlar (makbuz_no, tarih, islem_tipi, gonderen_tip, gonderen_id, gonderen_adi, alan_tip, alan_id, alan_adi, toplam_miktar, aciklama, yapan_kullanici_id, yapan_kullanici_adi) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id",
                   (makbuz_no, tarih, transfer_data['islem_tipi'], transfer_data['gonderen_tip'], transfer_data['gonderen_id'], transfer_data['gonderen_adi'], transfer_data['alan_tip'], transfer_data['alan_id'], transfer_data['alan_adi'], transfer_data['toplam_miktar'], transfer_data['aciklama'], transfer_data['yapan_id'], transfer_data['yapan_adi']))
    makbuz_id = cursor.fetchone()[0]
    for detay in transfer_data['detaylar']:
        cursor.execute("INSERT INTO makbuz_detaylari (makbuz_id, palet_tipi_id, stok_kodu, palet_adi, miktar) VALUES (%s, %s, %s, %s, %s)", (makbuz_id, detay['palet_tipi_id'], detay['stok_kodu'], detay['palet_adi'], detay['miktar']))
    conn.commit()
    cursor.close()
    conn.close()
    return makbuz_no


def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'hata': 'Token gerekli'}), 401
        try:
            token = token.replace('Bearer ', '')
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            current_user = data
        except:
            return jsonify({'hata': 'Geçersiz token'}), 401
        return f(current_user, *args, **kwargs)
    return decorated


@app.route('/')
def index():
    return send_from_directory('static', 'index.html')


@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, kullanici_adi, tip, ad_soyad, COALESCE(kisitlamalar, '') FROM kullanicilar WHERE kullanici_adi = %s AND sifre = %s AND COALESCE(aktif,1) = 1", (data.get('kullanici_adi'), hash_sifre(data.get('sifre'))))
    kullanici = cursor.fetchone()
    cursor.close()
    conn.close()
    if kullanici:
        token = jwt.encode({'id': kullanici[0], 'kullanici_adi': kullanici[1], 'tip': kullanici[2], 'ad_soyad': kullanici[3], 'kisitlamalar': kullanici[4], 'exp': datetime.utcnow() + timedelta(hours=24)}, app.config['SECRET_KEY'], algorithm='HS256')
        return jsonify({'success': True, 'token': token, 'kullanici': {'id': kullanici[0], 'kullanici_adi': kullanici[1], 'tip': kullanici[2], 'ad_soyad': kullanici[3], 'kisitlamalar': kullanici[4]}})
    return jsonify({'success': False, 'hata': 'Hatalı kullanıcı adı veya şifre'}), 401


@app.route('/api/kullanici_listesi', methods=['GET'])
@token_required
def get_kullanici_listesi(current_user):
    if current_user['tip'] != 'DEPOCU':
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, kullanici_adi, ad_soyad, COALESCE(kisitlamalar, ''), tip FROM kullanicilar WHERE tip IN ('DAGITICI', 'FORKLIFT') AND COALESCE(aktif,1) = 1 ORDER BY tip, ad_soyad")
    sonuc = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify([{'id': k[0], 'kullanici_adi': k[1], 'ad_soyad': k[2], 'kisitlamalar': k[3], 'tip': k[4]} for k in sonuc])


@app.route('/api/dagitici_listesi', methods=['GET'])
@token_required
def get_dagitici_listesi(current_user):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, kullanici_adi, ad_soyad FROM kullanicilar WHERE tip = 'DAGITICI' ORDER BY ad_soyad")
    sonuc = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify([{'id': d[0], 'kullanici_adi': d[1], 'ad_soyad': d[2]} for d in sonuc])


@app.route('/api/dagitici_ekle', methods=['POST'])
@token_required
def dagitici_ekle(current_user):
    if current_user['tip'] != 'DEPOCU':
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json()
    kullanici_adi, ad_soyad, sifre = data.get('kullanici_adi'), data.get('ad_soyad'), data.get('sifre')
    kisitlamalar = data.get('kisitlamalar', '')
    if not kullanici_adi or not ad_soyad or not sifre:
        return jsonify({'hata': 'Tüm alanlar gerekli'}), 400
    if len(sifre) < 4:
        return jsonify({'hata': 'Şifre en az 4 karakter olmalı'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO kullanicilar (kullanici_adi, sifre, tip, ad_soyad, kisitlamalar) VALUES (%s, %s, %s, %s, %s) RETURNING id", (kullanici_adi, hash_sifre(sifre), 'DAGITICI', ad_soyad, kisitlamalar))
        dagitici_id = cursor.fetchone()[0]
        cursor.execute("SELECT id FROM palet_tipleri")
        for palet in cursor.fetchall():
            cursor.execute("INSERT INTO stoklar (stok_sahibi_tip, stok_sahibi_id, palet_tipi_id, miktar) SELECT %s, %s, %s, 0 WHERE NOT EXISTS (SELECT 1 FROM stoklar WHERE stok_sahibi_tip = %s AND stok_sahibi_id = %s AND palet_tipi_id = %s)", ('DAGITICI', dagitici_id, palet[0], 'DAGITICI', dagitici_id, palet[0]))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'id': dagitici_id, 'mesaj': 'Dağıtıcı eklendi'})
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'hata': str(e)}), 400


@app.route('/api/forklift_ekle', methods=['POST'])
@token_required
def forklift_ekle(current_user):
    if current_user['tip'] != 'DEPOCU':
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json()
    kullanici_adi, ad_soyad, sifre, kisitlamalar = data.get('kullanici_adi'), data.get('ad_soyad'), data.get('sifre'), data.get('kisitlamalar', '')
    if not kullanici_adi or not ad_soyad or not sifre:
        return jsonify({'hata': 'Tüm alanlar gerekli'}), 400
    if len(sifre) < 4:
        return jsonify({'hata': 'Şifre en az 4 karakter olmalı'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO kullanicilar (kullanici_adi, sifre, tip, ad_soyad, kisitlamalar) VALUES (%s, %s, %s, %s, %s) RETURNING id", (kullanici_adi, hash_sifre(sifre), 'FORKLIFT', ad_soyad, kisitlamalar))
        forklift_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'id': forklift_id, 'mesaj': 'Forklift operatörü eklendi'})
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'hata': str(e)}), 400


@app.route('/api/kullanici_duzenle', methods=['PUT'])
@token_required
def kullanici_duzenle(current_user):
    if current_user['tip'] != 'DEPOCU':
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json()
    uid, kullanici_adi, ad_soyad, sifre, kisitlamalar = data.get('id'), data.get('kullanici_adi'), data.get('ad_soyad'), data.get('sifre'), data.get('kisitlamalar', '')
    if not uid or not kullanici_adi or not ad_soyad:
        return jsonify({'hata': 'id, kullanici_adi ve ad_soyad gerekli'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if sifre:
            db_execute(cursor, conn, "UPDATE kullanicilar SET kullanici_adi=%s, ad_soyad=%s, sifre=%s, kisitlamalar=%s WHERE id=%s AND tip IN ('DAGITICI','FORKLIFT')", (kullanici_adi, ad_soyad, hash_sifre(sifre), kisitlamalar, uid))
        else:
            db_execute(cursor, conn, "UPDATE kullanicilar SET kullanici_adi=%s, ad_soyad=%s, kisitlamalar=%s WHERE id=%s AND tip IN ('DAGITICI','FORKLIFT')", (kullanici_adi, ad_soyad, kisitlamalar, uid))
        if cursor.rowcount == 0:
            conn.rollback()
            cursor.close()
            conn.close()
            return jsonify({'hata': 'Kullanıcı bulunamadı veya güncellenemedi (id hatalı olabilir)'}), 404
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'mesaj': 'Kullanıcı güncellendi'})
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'hata': str(e)}), 400


@app.route('/api/kullanici_sil', methods=['DELETE'])
@token_required
def kullanici_sil(current_user):
    if current_user['tip'] != 'DEPOCU':
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    uid = request.args.get('id', type=int)
    if not uid:
        return jsonify({'hata': 'id gerekli'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Gerçekten silmek yerine pasife al - geçmiş hareketler korunur
        if _DB_TYPE == 'sqlite':
            cursor.execute("UPDATE kullanicilar SET aktif=0, kullanici_adi=kullanici_adi||'_silindi_'||CAST(id AS TEXT) WHERE id=? AND tip IN ('DAGITICI','FORKLIFT')", (uid,))
        else:
            cursor.execute("UPDATE kullanicilar SET aktif=0, kullanici_adi=kullanici_adi||'_silindi_'||id::text WHERE id=%s AND tip IN ('DAGITICI','FORKLIFT')", (uid,))
        if cursor.rowcount == 0:
            cursor.close()
            conn.close()
            return jsonify({'hata': 'Kullanıcı bulunamadı'}), 404
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'mesaj': 'Kullanıcı silindi'})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'hata': str(e)}), 400




@app.route('/api/tum_musteriler', methods=['GET'])
@token_required
def get_tum_musteriler(current_user):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, musteri_kodu, musteri_adi, tabela_adi FROM musteriler ORDER BY musteri_adi")
    sonuc = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify([{'id': m[0], 'musteri_kodu': m[1], 'musteri_adi': m[2], 'tabela_adi': m[3]} for m in sonuc])


@app.route('/api/musteri_ekle', methods=['POST'])
@token_required
def musteri_ekle(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json()
    musteri_kodu, musteri_adi, tabela_adi = data.get('musteri_kodu'), data.get('musteri_adi'), data.get('tabela_adi')
    if not musteri_kodu or not musteri_adi or not tabela_adi:
        return jsonify({'hata': 'Tüm alanlar gerekli'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO musteriler (musteri_kodu, musteri_adi, tabela_adi) VALUES (%s, %s, %s) RETURNING id", (musteri_kodu, musteri_adi, tabela_adi))
        musteri_id = cursor.fetchone()[0]
        cursor.execute("SELECT id FROM palet_tipleri")
        for palet in cursor.fetchall():
            cursor.execute("INSERT INTO stoklar (stok_sahibi_tip, stok_sahibi_id, palet_tipi_id, miktar) SELECT %s, %s, %s, 0 WHERE NOT EXISTS (SELECT 1 FROM stoklar WHERE stok_sahibi_tip = %s AND stok_sahibi_id = %s AND palet_tipi_id = %s)", ('MUSTERI', musteri_id, palet[0], 'MUSTERI', musteri_id, palet[0]))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'id': musteri_id, 'mesaj': 'Müşteri eklendi'})
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'hata': str(e)}), 400


@app.route('/api/palet_tipleri', methods=['GET'])
@token_required
def get_palet_tipleri(current_user):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, stok_kodu, palet_adi FROM palet_tipleri")
    sonuc = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify([{'id': p[0], 'stok_kodu': p[1], 'palet_adi': p[2]} for p in sonuc])


@app.route('/api/stok', methods=['GET'])
@token_required
def get_stok(current_user):
    tip, kimlik = request.args.get('tip'), request.args.get('id', type=int)
    if not tip or kimlik is None:
        return jsonify({'hata': 'tip ve id parametreleri gerekli'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT pt.id, pt.stok_kodu, pt.palet_adi, COALESCE(s.miktar, 0) FROM palet_tipleri pt LEFT JOIN stoklar s ON pt.id = s.palet_tipi_id AND s.stok_sahibi_tip = %s AND s.stok_sahibi_id = %s ORDER BY pt.id", (tip, kimlik))
    sonuc = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify([{'palet_id': p[0], 'stok_kodu': p[1], 'palet_adi': p[2], 'miktar': p[3]} for p in sonuc])


@app.route('/api/musteri_stoklari', methods=['GET'])
@token_required
def get_musteri_stoklari(current_user):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT m.id, m.musteri_kodu, m.musteri_adi, m.tabela_adi, 
               pt.stok_kodu, pt.palet_adi, s.miktar 
        FROM stoklar s 
        JOIN musteriler m ON s.stok_sahibi_id = m.id 
        JOIN palet_tipleri pt ON s.palet_tipi_id = pt.id 
        WHERE s.stok_sahibi_tip = 'MUSTERI' AND s.miktar > 0 
        ORDER BY m.musteri_adi
    ''')
    sonuc = cursor.fetchall()
    cursor.close()
    conn.close()
    musteriler = {}
    for row in sonuc:
        m_id = row[0]
        if m_id not in musteriler:
            musteriler[m_id] = {
                'id': m_id,
                'musteri_kodu': row[1],
                'musteri_adi': row[2],
                'tabela_adi': row[3],
                'stoklar': []
            }
        musteriler[m_id]['stoklar'].append({
            'stok_kodu': row[4],
            'palet_adi': row[5],
            'miktar': row[6]
        })
    return jsonify(list(musteriler.values()))


@app.route('/api/transfer', methods=['POST'])
@token_required
def transfer_yap(current_user):
    data = request.get_json()
    hareket_tipi, palet_tipi_id, miktar, alici_id = data.get('hareket_tipi'), data.get('palet_tipi_id'), data.get('miktar'), data.get('alici_id')
    if not hareket_tipi or not palet_tipi_id or not miktar:
        return jsonify({'hata': 'Eksik parametreler'}), 400
    if miktar <= 0:
        return jsonify({'hata': 'Miktar pozitif olmalı'}), 400
    kullanici_id, kullanici_tip, kullanici_adi, kullanici_kadi = current_user['id'], current_user['tip'], current_user['ad_soyad'], current_user['kullanici_adi']
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, stok_kodu, palet_adi FROM palet_tipleri WHERE id = %s", (palet_tipi_id,))
    palet = cursor.fetchone()
    if not palet:
        cursor.close()
        conn.close()
        return jsonify({'hata': 'Geçersiz palet tipi'}), 400
    transfer_data = {'islem_tipi': hareket_tipi, 'yapan_id': kullanici_id, 'yapan_adi': f"{kullanici_adi} ({kullanici_kadi})", 'detaylar': [], 'toplam_miktar': miktar}
    if kullanici_tip in ('DEPOCU', 'FORKLIFT'):
        if hareket_tipi == 'DEPO_DAGITICI':
            gonderen_tip, gonderen_id, gonderen_adi = SAHIP_TIP_DEPO, 0, "DEPO"
            alan_tip, alan_id = None, alici_id
            cursor.execute("SELECT musteri_kodu, musteri_adi FROM musteriler WHERE id = %s", (alici_id,))
            musteri = cursor.fetchone()
            if musteri:
                alan_tip = SAHIP_TIP_MUSTERI
                alan_adi = f"{musteri[0]} - {musteri[1]}"
                aciklama = f"{palet[2]} - {miktar} adet {musteri[1]} müşterisine (Depodan Doğrudan) verildi"
            else:
                cursor.execute("SELECT ad_soyad, kullanici_adi FROM kullanicilar WHERE id = %s AND tip = 'DAGITICI'", (alici_id,))
                dagitici = cursor.fetchone()
                if not dagitici:
                    cursor.close()
                    conn.close()
                    return jsonify({'hata': 'Geçersiz alıcı ID (Müşteri veya Dağıtıcı bulunamadı)'}), 400
                alan_tip = SAHIP_TIP_DAGITICI
                alan_adi = f"{dagitici[0]} ({dagitici[1]})"
                aciklama = f"{palet[2]} - {miktar} adet {dagitici[0]} dağıtıcısına transfer edildi"
        elif hareket_tipi == 'DAGITICI_DEPO':
            alan_tip, alan_id, alan_adi = SAHIP_TIP_DEPO, 0, "DEPO"
            gonderen_tip, gonderen_id = None, alici_id
            cursor.execute("SELECT musteri_kodu, musteri_adi FROM musteriler WHERE id = %s", (alici_id,))
            musteri = cursor.fetchone()
            if musteri:
                gonderen_tip = SAHIP_TIP_MUSTERI
                gonderen_adi = f"{musteri[0]} - {musteri[1]}"
                aciklama = f"{palet[2]} - {miktar} adet {musteri[1]} müşterisinden (Depoya Doğrudan) iade alındı"
            else:
                cursor.execute("SELECT ad_soyad, kullanici_adi FROM kullanicilar WHERE id = %s AND tip = 'DAGITICI'", (alici_id,))
                dagitici = cursor.fetchone()
                if not dagitici:
                    cursor.close()
                    conn.close()
                    return jsonify({'hata': 'Geçersiz gönderen ID'}), 400
                gonderen_tip = SAHIP_TIP_DAGITICI
                gonderen_adi = f"{dagitici[0]} ({dagitici[1]})"
                aciklama = f"{palet[2]} - {miktar} adet {dagitici[0]} dağıtıcısından iade alındı"
    elif kullanici_tip == 'DAGITICI':
        if hareket_tipi == 'DAGITICI_MUSTERI':
            gonderen_tip, gonderen_id, gonderen_adi = SAHIP_TIP_DAGITICI, kullanici_id, f"{kullanici_adi} ({kullanici_kadi})"
            alan_tip, alan_id = SAHIP_TIP_MUSTERI, alici_id
            cursor.execute("SELECT musteri_kodu, musteri_adi FROM musteriler WHERE id = %s", (alici_id,))
            musteri = cursor.fetchone()
            if not musteri:
                cursor.close()
                conn.close()
                return jsonify({'hata': 'Geçersiz müşteri ID'}), 400
            alan_adi = f"{musteri[0]} - {musteri[1]}"
            aciklama = f"{palet[2]} - {miktar} adet {musteri[1]} müşterisine verildi"
        elif hareket_tipi == 'MUSTERI_DAGITICI':
            gonderen_tip, gonderen_id = SAHIP_TIP_MUSTERI, alici_id
            alan_tip, alan_id, alan_adi = SAHIP_TIP_DAGITICI, kullanici_id, f"{kullanici_adi} ({kullanici_kadi})"
            cursor.execute("SELECT musteri_kodu, musteri_adi FROM musteriler WHERE id = %s", (alici_id,))
            musteri = cursor.fetchone()
            if not musteri:
                cursor.close()
                conn.close()
                return jsonify({'hata': 'Geçersiz müşteri ID'}), 400
            gonderen_adi = f"{musteri[0]} - {musteri[1]}"
            aciklama = f"{palet[2]} - {miktar} adet {musteri[1]} müşterisinden iade alındı"
        else:
            cursor.close()
            conn.close()
            return jsonify({'hata': 'Geçersiz hareket tipi'}), 400
    else:
        cursor.close()
        conn.close()
        return jsonify({'hata': 'Yetkisiz kullanıcı'}), 403
    transfer_data.update({'gonderen_tip': gonderen_tip, 'gonderen_id': gonderen_id, 'gonderen_adi': gonderen_adi, 'alan_tip': alan_tip, 'alan_id': alan_id, 'alan_adi': alan_adi, 'aciklama': aciklama})
    transfer_data['detaylar'].append({'palet_tipi_id': palet_tipi_id, 'stok_kodu': palet[1], 'palet_adi': palet[2], 'miktar': miktar})
    mevcut_gonderen = stok_miktari_getir(gonderen_tip, gonderen_id, palet_tipi_id)
    if mevcut_gonderen < miktar:
        cursor.close()
        conn.close()
        return jsonify({'hata': f'Yetersiz stok! Mevcut: {mevcut_gonderen}'}), 400
    basarili, hata = stok_guncelle(gonderen_tip, gonderen_id, palet_tipi_id, -miktar)
    if not basarili:
        cursor.close()
        conn.close()
        return jsonify({'hata': hata}), 400
    basarili, hata = stok_guncelle(alan_tip, alan_id, palet_tipi_id, +miktar)
    if not basarili:
        stok_guncelle(gonderen_tip, gonderen_id, palet_tipi_id, +miktar)
        cursor.close()
        conn.close()
        return jsonify({'hata': hata}), 400
    makbuz_no = makbuz_kaydet(transfer_data)
    hareket_kaydet(kullanici_id, hareket_tipi, gonderen_tip, gonderen_id, alan_tip, alan_id, palet_tipi_id, miktar, aciklama, makbuz_no)
    cursor.close()
    conn.close()
    return jsonify({'success': True, 'mesaj': f'Transfer başarılı! {miktar} adet {palet[2]} transfer edildi.', 'makbuz_no': makbuz_no})


@app.route('/api/makbuz/<makbuz_no>', methods=['GET'])
@token_required
def makbuz_goster(current_user, makbuz_no):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM makbuzlar WHERE makbuz_no = %s", (makbuz_no,))
    makbuz = cursor.fetchone()
    if not makbuz:
        cursor.close()
        conn.close()
        return jsonify({'hata': 'Makbuz bulunamadı'}), 404
    cursor.execute("SELECT stok_kodu, palet_adi, miktar FROM makbuz_detaylari WHERE makbuz_id = %s", (makbuz[0],))
    detaylar = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify({'makbuz_no': makbuz[1], 'tarih': makbuz[2], 'islem_tipi': makbuz[3], 'gonderen_adi': makbuz[6], 'alan_adi': makbuz[9], 'toplam_miktar': makbuz[10], 'aciklama': makbuz[11], 'yapan_adi': makbuz[13], 'detaylar': [{'stok_kodu': d[0], 'palet_adi': d[1], 'miktar': d[2]} for d in detaylar]})


@app.route('/api/makbuz/pdf/<makbuz_no>', methods=['GET'])
@token_required
def makbuz_pdf(current_user, makbuz_no):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM makbuzlar WHERE makbuz_no = %s", (makbuz_no,))
    makbuz = cursor.fetchone()
    if not makbuz:
        cursor.close()
        conn.close()
        return jsonify({'hata': 'Makbuz bulunamadı'}), 404
    cursor.execute("SELECT stok_kodu, palet_adi, miktar FROM makbuz_detaylari WHERE makbuz_id = %s", (makbuz[0],))
    detaylar = cursor.fetchall()
    cursor.close()
    conn.close()
    font_path = os.path.join(os.path.dirname(__file__), 'arial.ttf')
    try:
        pdfmetrics.registerFont(TTFont('TurkceFont', font_path))
        aktif_font = 'TurkceFont'
    except Exception as e:
        aktif_font = 'Helvetica'
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    styles['Normal'].fontName = aktif_font
    styles['Heading1'].fontName = aktif_font
    styles['Heading2'].fontName = aktif_font
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontName=aktif_font, fontSize=16, alignment=1, spaceAfter=20)
    normal_style = styles['Normal']
    story = []
    story.append(Paragraph("ULUDAĞ İÇECEK", title_style))
    story.append(Paragraph("KONYA BÖLGE DEPO", title_style))
    story.append(Paragraph("PALET İŞLEM MAKBUZU", title_style))
    story.append(Spacer(1, 20))
    tip_text = {'DEPO_DAGITICI': 'Depo → Dağıtıcı', 'DAGITICI_MUSTERI': 'Dağıtıcı → Müşteri', 'MUSTERI_DAGITICI': 'Müşteri → Dağıtıcı', 'DAGITICI_DEPO': 'Dağıtıcı → Depo'}.get(makbuz[3], makbuz[3])
    data = [
        ['Makbuz No:', makbuz[1]], 
        ['Tarih:', makbuz[2]], 
        ['İşlem Türü:', tip_text], 
        ['Teslim Eden:', Paragraph(makbuz[6], normal_style)], 
        ['Teslim Alan:', Paragraph(makbuz[9], normal_style)], 
        ['Toplam Miktar:', str(makbuz[10])], 
        ['Açıklama:', Paragraph(makbuz[11] or "", normal_style)], 
        ['İşlemi Yapan:', makbuz[13]]
    ]
    table = Table(data, colWidths=[100, 300])
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, -1), aktif_font),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    story.append(table)
    story.append(Spacer(1, 20))
    detay_data = [['Stok Kodu', 'Palet Adı', 'Miktar']] + [[d[0], Paragraph(d[1], normal_style), str(d[2])] for d in detaylar]
    detay_table = Table(detay_data, colWidths=[100, 200, 80])
    detay_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), aktif_font),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    story.append(detay_table)
    story.append(Spacer(1, 30))
    story.append(Paragraph("_________________________", normal_style))
    story.append(Paragraph("Yetkili İmza", normal_style))
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'makbuz_{makbuz_no}.pdf')


@app.route('/api/hareketler', methods=['GET'])
@token_required
def get_hareketler(current_user):
    limit = request.args.get('limit', 50, type=int)
    conn = get_db_connection()
    cursor = conn.cursor()
    if current_user['tip'] in ('DEPOCU', 'FORKLIFT'):
        cursor.execute("SELECT h.tarih, u.kullanici_adi, u.ad_soyad, h.hareket_tipi, pt.stok_kodu, pt.palet_adi, h.miktar, h.aciklama, h.makbuz_no FROM hareketler h JOIN kullanicilar u ON h.yapan_kullanici_id = u.id JOIN palet_tipleri pt ON h.palet_tipi_id = pt.id ORDER BY h.tarih DESC LIMIT %s", (limit,))
    else:
        cursor.execute("SELECT h.tarih, u.kullanici_adi, u.ad_soyad, h.hareket_tipi, pt.stok_kodu, pt.palet_adi, h.miktar, h.aciklama, h.makbuz_no FROM hareketler h JOIN kullanicilar u ON h.yapan_kullanici_id = u.id JOIN palet_tipleri pt ON h.palet_tipi_id = pt.id WHERE h.gonderen_id = %s OR h.alan_id = %s ORDER BY h.tarih DESC LIMIT %s", (current_user['id'], current_user['id'], limit))
    sonuc = cursor.fetchall()
    cursor.close()
    conn.close()
    hareketler = []
    for h in sonuc:
        tip_text = {'DEPO_DAGITICI': 'Depo→Dağıtıcı', 'DAGITICI_MUSTERI': 'Dağıtıcı→Müşteri', 'MUSTERI_DAGITICI': 'Müşteri→Dağıtıcı', 'DAGITICI_DEPO': 'Dağıtıcı→Depo', 'DEPO_STOK_HAREKET': 'Depo Stok Hareketi'}.get(h[3], h[3])
        hareketler.append({'tarih': h[0], 'yapan': f"{h[2]} ({h[1]})", 'islem_tipi': tip_text, 'stok_kodu': h[4], 'palet_adi': h[5], 'miktar': h[6], 'aciklama': h[7], 'makbuz_no': h[8]})
    return jsonify(hareketler)


@app.route('/api/hareketler_filtreli', methods=['POST'])
@token_required
def get_hareketler_filtreli(current_user):
    data = request.get_json()
    dagitici_id, palet_tipi_id, baslangic, bitis = data.get('dagitici_id'), data.get('palet_tipi_id'), data.get('baslangic_tarihi'), data.get('bitis_tarihi')
    limit = data.get('limit', 100)
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """SELECT h.tarih, u.kullanici_adi, u.ad_soyad, h.hareket_tipi, pt.stok_kodu, pt.palet_adi, h.miktar, h.aciklama, h.makbuz_no,
               CASE WHEN h.gonderen_tip = 'DAGITICI' THEN (SELECT ad_soyad FROM kullanicilar WHERE id = h.gonderen_id)
                    WHEN h.alan_tip = 'DAGITICI' THEN (SELECT ad_soyad FROM kullanicilar WHERE id = h.alan_id) ELSE NULL END as ilgili_dagitici
               FROM hareketler h JOIN kullanicilar u ON h.yapan_kullanici_id = u.id JOIN palet_tipleri pt ON h.palet_tipi_id = pt.id WHERE 1=1"""
    params = []
    if dagitici_id:
        query += " AND (h.gonderen_id = %s OR h.alan_id = %s)"
        params.extend([dagitici_id, dagitici_id])
    if palet_tipi_id:
        query += " AND h.palet_tipi_id = %s"
        params.append(palet_tipi_id)
    if baslangic and bitis:
        query += " AND DATE(h.tarih) BETWEEN %s AND %s"
        params.extend([baslangic, bitis])
    query += " ORDER BY h.tarih DESC LIMIT %s"
    params.append(limit)
    cursor.execute(query, params)
    sonuc = cursor.fetchall()
    cursor.close()
    conn.close()
    hareketler = []
    for h in sonuc:
        tip_text = {'DEPO_DAGITICI': 'Depo→Dağıtıcı', 'DAGITICI_MUSTERI': 'Dağıtıcı→Müşteri', 'MUSTERI_DAGITICI': 'Müşteri→Dağıtıcı', 'DAGITICI_DEPO': 'Dağıtıcı→Depo', 'DEPO_STOK_HAREKET': 'Depo Stok Hareketi'}.get(h[3], h[3])
        hareketler.append({'tarih': h[0], 'yapan': f"{h[2]} ({h[1]})", 'islem_tipi': tip_text, 'stok_kodu': h[4], 'palet_adi': h[5], 'miktar': h[6], 'aciklama': h[7], 'makbuz_no': h[8], 'ilgili_dagitici': h[9] or '-'})
    return jsonify(hareketler)


@app.route('/api/depo_stok_hareket', methods=['POST'])
@token_required
def depo_stok_hareket(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json()
    palet_tipi_id, miktar, islem_tipi, aciklama = data.get('palet_tipi_id'), data.get('miktar'), data.get('islem_tipi'), data.get('aciklama', '')
    if not palet_tipi_id or not miktar or not islem_tipi:
        return jsonify({'hata': 'Eksik parametreler'}), 400
    if miktar <= 0:
        return jsonify({'hata': 'Miktar pozitif olmalı'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT stok_kodu, palet_adi FROM palet_tipleri WHERE id = %s", (palet_tipi_id,))
    palet = cursor.fetchone()
    if not palet:
        cursor.close()
        conn.close()
        return jsonify({'hata': 'Geçersiz palet tipi'}), 400
    mevcut = stok_miktari_getir(SAHIP_TIP_DEPO, 0, palet_tipi_id)
    if islem_tipi == 'azalt':
        if mevcut < miktar:
            cursor.close()
            conn.close()
            return jsonify({'hata': f'Yetersiz stok! Mevcut: {mevcut}'}), 400
        yeni_miktar, degisim = mevcut - miktar, -miktar
        hareket_aciklama = f"STOK AZALTMA: {miktar} adet {palet[1]} azaltıldı. Sebep: {aciklama}"
    else:
        yeni_miktar, degisim = mevcut + miktar, +miktar
        hareket_aciklama = f"STOK ARTTIRMA: {miktar} adet {palet[1]} eklendi. Sebep: {aciklama}"
    basarili, hata = stok_guncelle(SAHIP_TIP_DEPO, 0, palet_tipi_id, degisim)
    if not basarili:
        cursor.close()
        conn.close()
        return jsonify({'hata': hata}), 400
    hareket_kaydet(current_user['id'], "DEPO_STOK_HAREKET", SAHIP_TIP_DEPO, 0, SAHIP_TIP_DEPO, 0, palet_tipi_id, miktar, hareket_aciklama)
    cursor.close()
    conn.close()
    return jsonify({'success': True, 'mesaj': f'{palet[1]} stoğu güncellendi. Yeni miktar: {yeni_miktar}', 'yeni_miktar': yeni_miktar})


@app.route('/api/rapor/istatistikler', methods=['GET'])
@token_required
def rapor_istatistikler(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # 1. Dağıtıcı performansı: verilen, toplanan, oran
        cursor.execute("""
            SELECT u.ad_soyad,
                   SUM(CASE WHEN h.hareket_tipi = 'DAGITICI_MUSTERI' THEN h.miktar ELSE 0 END) as verilen,
                   SUM(CASE WHEN h.hareket_tipi = 'MUSTERI_DAGITICI' THEN h.miktar ELSE 0 END) as toplanan
            FROM kullanicilar u
            LEFT JOIN hareketler h ON h.yapan_kullanici_id = u.id
            WHERE u.tip = 'DAGITICI' AND COALESCE(u.aktif,1)=1
            GROUP BY u.id, u.ad_soyad
            ORDER BY verilen DESC
        """)
        dagitici_perf = []
        for r in cursor.fetchall():
            v, t = int(r[1] or 0), int(r[2] or 0)
            oran = round(t / v * 100, 1) if v > 0 else 0
            dagitici_perf.append({'ad': r[0], 'verilen': v, 'toplanan': t, 'oran': oran, 'bekleyen': v - t})

        # 2. En çok palet bekleyen müşteriler (top 15)
        cursor.execute("""
            SELECT m.musteri_kodu, m.musteri_adi, SUM(s.miktar) as toplam
            FROM stoklar s JOIN musteriler m ON s.stok_sahibi_id = m.id
            WHERE s.stok_sahibi_tip = 'MUSTERI' AND s.miktar > 0
            GROUP BY m.id, m.musteri_kodu, m.musteri_adi
            ORDER BY toplam DESC LIMIT 15
        """)
        musteri_stok = [{'kod': r[0], 'ad': r[1], 'miktar': int(r[2])} for r in cursor.fetchall()]

        # 3. Genel stok özeti
        cursor.execute("SELECT SUM(miktar) FROM stoklar WHERE stok_sahibi_tip='DEPO'")
        depo = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT SUM(miktar) FROM stoklar WHERE stok_sahibi_tip='DAGITICI'")
        dagitici_stok = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT SUM(miktar) FROM stoklar WHERE stok_sahibi_tip='MUSTERI'")
        musteri_stok_toplam = int(cursor.fetchone()[0] or 0)

        # 4. Palet tipi bazında dağılım
        cursor.execute("""
            SELECT pt.palet_adi,
                   SUM(CASE WHEN s.stok_sahibi_tip='DEPO' THEN s.miktar ELSE 0 END) as depo_adet,
                   SUM(CASE WHEN s.stok_sahibi_tip='DAGITICI' THEN s.miktar ELSE 0 END) as dagitici_adet,
                   SUM(CASE WHEN s.stok_sahibi_tip='MUSTERI' THEN s.miktar ELSE 0 END) as musteri_adet,
                   SUM(s.miktar) as toplam
            FROM stoklar s JOIN palet_tipleri pt ON s.palet_tipi_id = pt.id
            GROUP BY pt.id, pt.palet_adi ORDER BY toplam DESC
        """)
        palet_dagilim = [{'ad': r[0], 'depo': int(r[1] or 0), 'dagitici': int(r[2] or 0), 'musteri': int(r[3] or 0), 'toplam': int(r[4] or 0)} for r in cursor.fetchall()]

        # 5. Son 30 gün günlük hareket trendi - tarih TEXT olarak saklandığı için SUBSTR ile kes
        sinir_tarih = (get_now() - timedelta(days=30)).strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT SUBSTR(tarih, 1, 10) as gun, SUM(miktar)
            FROM hareketler
            WHERE hareket_tipi = 'DAGITICI_MUSTERI'
              AND tarih >= %s
            GROUP BY gun ORDER BY gun
        """, (sinir_tarih,))
        trend = [{'gun': str(r[0]), 'miktar': int(r[1] or 0)} for r in cursor.fetchall()]

    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'hata': f'Sorgu hatası: {str(e)}'}), 500

    cursor.close()
    conn.close()
    return jsonify({
        'dagitici_perf': dagitici_perf,
        'musteri_stok': musteri_stok,
        'stok_ozet': {'depo': depo, 'dagitici': dagitici_stok, 'musteri': musteri_stok_toplam, 'toplam': depo + dagitici_stok + musteri_stok_toplam},
        'palet_dagilim': palet_dagilim,
        'trend': trend
    })


@app.route('/api/rapor/export', methods=['POST'])
@token_required
def rapor_export(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json()
    rapor_tipi, baslangic, bitis = data.get('rapor_tipi'), data.get('baslangic_tarihi'), data.get('bitis_tarihi')
    conn = get_db_connection()
    cursor = conn.cursor()
    workbook = openpyxl.Workbook()
    if rapor_tipi == 'hareketler':
        sheet = workbook.active
        sheet.title = "Hareketler"
        headers = ['Tarih', 'Yapan Kullanıcı', 'İşlem Tipi', 'Stok Kodu', 'Palet Adı', 'Miktar', 'Makbuz No', 'Açıklama']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        if baslangic and bitis:
            cursor.execute("SELECT h.tarih, u.ad_soyad, h.hareket_tipi, pt.stok_kodu, pt.palet_adi, h.miktar, h.makbuz_no, h.aciklama FROM hareketler h JOIN kullanicilar u ON h.yapan_kullanici_id = u.id JOIN palet_tipleri pt ON h.palet_tipi_id = pt.id WHERE DATE(h.tarih) BETWEEN %s AND %s ORDER BY h.tarih DESC", (baslangic, bitis))
        else:
            cursor.execute("SELECT h.tarih, u.ad_soyad, h.hareket_tipi, pt.stok_kodu, pt.palet_adi, h.miktar, h.makbuz_no, h.aciklama FROM hareketler h JOIN kullanicilar u ON h.yapan_kullanici_id = u.id JOIN palet_tipleri pt ON h.palet_tipi_id = pt.id ORDER BY h.tarih DESC LIMIT 1000")
        for row_idx, row in enumerate(cursor.fetchall(), 2):
            tip_text = {'DEPO_DAGITICI': 'Depo→Dağıtıcı', 'DAGITICI_MUSTERI': 'Dağıtıcı→Müşteri', 'MUSTERI_DAGITICI': 'Müşteri→Dağıtıcı', 'DAGITICI_DEPO': 'Dağıtıcı→Depo', 'DEPO_STOK_HAREKET': 'Depo Stok Hareketi'}.get(row[2], row[2])
            sheet.cell(row=row_idx, column=1, value=row[0])
            sheet.cell(row=row_idx, column=2, value=row[1])
            sheet.cell(row=row_idx, column=3, value=tip_text)
            sheet.cell(row=row_idx, column=4, value=row[3])
            sheet.cell(row=row_idx, column=5, value=row[4])
            sheet.cell(row=row_idx, column=6, value=row[5])
            sheet.cell(row=row_idx, column=7, value=row[6] or '-')
            sheet.cell(row=row_idx, column=8, value=row[7] or '')
        for col in range(1, 9):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    elif rapor_tipi == 'stoklar':
        sheet = workbook.active
        sheet.title = "Stoklar"
        headers = ['Stok Sahibi', 'Stok Kodu', 'Palet Adı', 'Miktar']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        row_idx = 2
        cursor.execute("SELECT pt.stok_kodu, pt.palet_adi, s.miktar FROM stoklar s JOIN palet_tipleri pt ON s.palet_tipi_id = pt.id WHERE s.stok_sahibi_tip = 'DEPO'")
        for row in cursor.fetchall():
            sheet.cell(row=row_idx, column=1, value="DEPO")
            sheet.cell(row=row_idx, column=2, value=row[0])
            sheet.cell(row=row_idx, column=3, value=row[1])
            sheet.cell(row=row_idx, column=4, value=row[2])
            row_idx += 1
        cursor.execute("SELECT u.ad_soyad, pt.stok_kodu, pt.palet_adi, s.miktar FROM stoklar s JOIN kullanicilar u ON s.stok_sahibi_id = u.id JOIN palet_tipleri pt ON s.palet_tipi_id = pt.id WHERE s.stok_sahibi_tip = 'DAGITICI' AND s.miktar > 0 ORDER BY u.ad_soyad")
        for row in cursor.fetchall():
            sheet.cell(row=row_idx, column=1, value=row[0])
            sheet.cell(row=row_idx, column=2, value=row[1])
            sheet.cell(row=row_idx, column=3, value=row[2])
            sheet.cell(row=row_idx, column=4, value=row[3])
            row_idx += 1
        cursor.execute("SELECT m.musteri_kodu, m.musteri_adi, pt.stok_kodu, pt.palet_adi, s.miktar FROM stoklar s JOIN musteriler m ON s.stok_sahibi_id = m.id JOIN palet_tipleri pt ON s.palet_tipi_id = pt.id WHERE s.stok_sahibi_tip = 'MUSTERI' AND s.miktar > 0 ORDER BY m.musteri_adi")
        for row in cursor.fetchall():
            sheet.cell(row=row_idx, column=1, value=f"{row[0]} - {row[1]}")
            sheet.cell(row=row_idx, column=2, value=row[2])
            sheet.cell(row=row_idx, column=3, value=row[3])
            sheet.cell(row=row_idx, column=4, value=row[4])
            row_idx += 1
        for col in range(1, 5):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    elif rapor_tipi == 'musteriler':
        sheet = workbook.active
        sheet.title = "Müşteriler"
        headers = ['Müşteri Kodu', 'Müşteri Adı', 'Tabela Adı']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        cursor.execute("SELECT musteri_kodu, musteri_adi, tabela_adi FROM musteriler ORDER BY musteri_adi")
        for row_idx, row in enumerate(cursor.fetchall(), 2):
            sheet.cell(row=row_idx, column=1, value=row[0])
            sheet.cell(row=row_idx, column=2, value=row[1])
            sheet.cell(row=row_idx, column=3, value=row[2])
        for col in range(1, 4):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    cursor.close()
    conn.close()
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'{rapor_tipi}_raporu.xlsx')


@app.route('/api/rapor/pdf', methods=['POST'])
@token_required
def rapor_pdf(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json()
    rapor_tipi = data.get('rapor_tipi', 'hareketler')
    baslangic = data.get('baslangic_tarihi')
    bitis = data.get('bitis_tarihi')
    conn = get_db_connection()
    cursor = conn.cursor()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story = []
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
    story.append(Paragraph("ULUDAĞ İÇECEK KONYA BÖLGE DEPO", title_style))
    story.append(Paragraph(f"{rapor_tipi.upper()} RAPORU", title_style))
    story.append(Spacer(1, 10))
    if baslangic and bitis:
        date_text = f"Tarih Aralığı: {baslangic} - {bitis}"
    else:
        date_text = f"Oluşturma Tarihi: {get_now().strftime('%d.%m.%Y %H:%M')}"
    story.append(Paragraph(date_text, styles['Normal']))
    story.append(Spacer(1, 20))
    if rapor_tipi == 'hareketler':
        if baslangic and bitis:
            cursor.execute('''SELECT h.tarih, u.ad_soyad, h.hareket_tipi, pt.stok_kodu, pt.palet_adi, h.miktar, h.aciklama, h.makbuz_no FROM hareketler h JOIN kullanicilar u ON h.yapan_kullanici_id = u.id JOIN palet_tipleri pt ON h.palet_tipi_id = pt.id WHERE DATE(h.tarih) BETWEEN %s AND %s ORDER BY h.tarih DESC''', (baslangic, bitis))
        else:
            cursor.execute('''SELECT h.tarih, u.ad_soyad, h.hareket_tipi, pt.stok_kodu, pt.palet_adi, h.miktar, h.aciklama, h.makbuz_no FROM hareketler h JOIN kullanicilar u ON h.yapan_kullanici_id = u.id JOIN palet_tipleri pt ON h.palet_tipi_id = pt.id ORDER BY h.tarih DESC LIMIT 500''')
        sonuc = cursor.fetchall()
        data = [['Tarih', 'Yapan', 'İşlem Tipi', 'Stok Kodu', 'Palet Adı', 'Miktar', 'Makbuz No', 'Açıklama']]
        for row in sonuc:
            tip_text = {'DEPO_DAGITICI': 'Depo→Dağıtıcı', 'DAGITICI_MUSTERI': 'Dağıtıcı→Müşteri', 'MUSTERI_DAGITICI': 'Müşteri→Dağıtıcı', 'DAGITICI_DEPO': 'Dağıtıcı→Depo', 'DEPO_STOK_HAREKET': 'Depo Stok Hareketi'}.get(row[2], row[2])
            data.append([row[0][:16], row[1], tip_text, row[3], row[4], str(row[5]), row[7] or '-', (row[6][:40] + '...') if row[6] and len(row[6]) > 40 else (row[6] or '')])
        table = Table(data, colWidths=[80, 70, 80, 60, 70, 40, 90, 100])
        table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 8), ('FONTSIZE', (0, 1), (-1, -1), 7), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        story.append(table)
    elif rapor_tipi == 'stoklar':
        cursor.execute('SELECT pt.stok_kodu, pt.palet_adi, s.miktar FROM stoklar s JOIN palet_tipleri pt ON s.palet_tipi_id = pt.id WHERE s.stok_sahibi_tip = %s', ('DEPO',))
        depo_stok = cursor.fetchall()
        story.append(Paragraph("📦 DEPO STOĞU", styles['Heading2']))
        data = [['Stok Kodu', 'Palet Adı', 'Miktar']] + [[r[0], r[1], str(r[2])] for r in depo_stok]
        table = Table(data, colWidths=[100, 200, 80])
        table.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        story.append(table)
        story.append(Spacer(1, 20))
        cursor.execute('''SELECT u.ad_soyad, pt.stok_kodu, pt.palet_adi, s.miktar FROM stoklar s JOIN kullanicilar u ON s.stok_sahibi_id = u.id JOIN palet_tipleri pt ON s.palet_tipi_id = pt.id WHERE s.stok_sahibi_tip = 'DAGITICI' AND s.miktar > 0 ORDER BY u.ad_soyad''')
        dagitici_stok = cursor.fetchall()
        story.append(Paragraph("🚚 DAĞITICI STOKLARI", styles['Heading2']))
        data = [['Dağıtıcı', 'Stok Kodu', 'Palet Adı', 'Miktar']] + [[r[0], r[1], r[2], str(r[3])] for r in dagitici_stok]
        table = Table(data, colWidths=[100, 80, 120, 60])
        table.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        story.append(table)
        story.append(Spacer(1, 20))
        cursor.execute('''SELECT m.musteri_kodu, m.musteri_adi, pt.stok_kodu, pt.palet_adi, s.miktar FROM stoklar s JOIN musteriler m ON s.stok_sahibi_id = m.id JOIN palet_tipleri pt ON s.palet_tipi_id = pt.id WHERE s.stok_sahibi_tip = 'MUSTERI' AND s.miktar > 0 ORDER BY m.musteri_adi LIMIT 100''')
        musteri_stok = cursor.fetchall()
        story.append(Paragraph("🏪 MÜŞTERİ STOKLARI", styles['Heading2']))
        data = [['Müşteri', 'Stok Kodu', 'Palet Adı', 'Miktar']] + [[f"{r[0]} - {r[1]}", r[2], r[3], str(r[4])] for r in musteri_stok]
        table = Table(data, colWidths=[150, 80, 100, 60])
        table.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        story.append(table)
    elif rapor_tipi == 'istatistik':
        cursor.execute('''SELECT u.ad_soyad, COUNT(*) as sayi FROM hareketler h JOIN kullanicilar u ON h.yapan_kullanici_id = u.id WHERE u.tip = 'DAGITICI' GROUP BY u.id ORDER BY sayi DESC LIMIT 10''')
        en_cok = cursor.fetchall()
        story.append(Paragraph("🏆 EN ÇOK TRANSFER YAPAN DAĞITICILAR", styles['Heading2']))
        data = [['Dağıtıcı', 'Transfer Sayısı']] + [[r[0], str(r[1])] for r in en_cok]
        table = Table(data, colWidths=[200, 100])
        table.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        story.append(table)
        story.append(Spacer(1, 20))
        cursor.execute('''SELECT pt.palet_adi, COUNT(*) as sayi, SUM(h.miktar) as toplam FROM hareketler h JOIN palet_tipleri pt ON h.palet_tipi_id = pt.id GROUP BY pt.id ORDER BY sayi DESC''')
        palet_kullanim = cursor.fetchall()
        story.append(Paragraph("📦 EN ÇOK KULLANILAN PALET TİPLERİ", styles['Heading2']))
        data = [['Palet Tipi', 'Kullanım Sayısı', 'Toplam Miktar']] + [[r[0], str(r[1]), str(r[2])] for r in palet_kullanim]
        table = Table(data, colWidths=[150, 100, 100])
        table.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        story.append(table)
    cursor.close()
    conn.close()
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'{rapor_tipi}_raporu_{get_now().strftime("%Y%m%d_%H%M%S")}.pdf')


@app.route('/api/musteri_excel_yukle', methods=['POST'])
@token_required
def musteri_excel_yukle(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    if 'file' not in request.files:
        return jsonify({'hata': 'Dosya bulunamadı'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'hata': 'Dosya seçilmedi'}), 400
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
    except Exception as e:
        return jsonify({'hata': f'Excel okunamadı: {str(e)}'}), 400
    headers = [str(sheet.cell(row=1, column=col).value or '').strip().upper() for col in range(1, sheet.max_column + 1)]
    required_columns = ['MÜŞTERİ KODU', 'MÜŞTERİ ADI', 'TABELA ADI']
    for col in required_columns:
        if col not in headers:
            return jsonify({'hata': f"'{col}' sütunu bulunamadı"}), 400
    col_indices = {h: i+1 for i, h in enumerate(headers)}
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT musteri_kodu, id FROM musteriler')
    mevcut_musteriler = {m[0]: m[1] for m in cursor.fetchall()}
    cursor.execute('SELECT id FROM palet_tipleri')
    paletler = [p[0] for p in cursor.fetchall()]
    eklenen = guncellenen = 0
    hatalar = []
    for row in range(2, sheet.max_row + 1):
        musteri_kodu = str(sheet.cell(row=row, column=col_indices['MÜŞTERİ KODU']).value or '').strip()
        musteri_adi = str(sheet.cell(row=row, column=col_indices['MÜŞTERİ ADI']).value or '').strip()
        tabela_adi = str(sheet.cell(row=row, column=col_indices['TABELA ADI']).value or '').strip()
        if not musteri_kodu or not musteri_adi or not tabela_adi:
            hatalar.append(f"Satır {row}: Boş alan var")
            continue
        if musteri_kodu in mevcut_musteriler:
            try:
                cursor.execute("UPDATE musteriler SET musteri_adi = %s, tabela_adi = %s WHERE musteri_kodu = %s", (musteri_adi, tabela_adi, musteri_kodu))
                guncellenen += 1
            except Exception as e:
                hatalar.append(f"Satır {row}: Güncelleme hatası - {str(e)}")
        else:
            try:
                cursor.execute("INSERT INTO musteriler (musteri_kodu, musteri_adi, tabela_adi) VALUES (%s, %s, %s) RETURNING id", (musteri_kodu, musteri_adi, tabela_adi))
                musteri_id = cursor.fetchone()[0]
                for palet_id in paletler:
                    cursor.execute("INSERT INTO stoklar (stok_sahibi_tip, stok_sahibi_id, palet_tipi_id, miktar) SELECT %s, %s, %s, 0 WHERE NOT EXISTS (SELECT 1 FROM stoklar WHERE stok_sahibi_tip = %s AND stok_sahibi_id = %s AND palet_tipi_id = %s)", ('MUSTERI', musteri_id, palet_id, 'MUSTERI', musteri_id, palet_id))
                eklenen += 1
                mevcut_musteriler[musteri_kodu] = musteri_id
            except Exception as e:
                hatalar.append(f"Satır {row}: Ekleme hatası - {str(e)}")
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({'success': True, 'eklenen': eklenen, 'guncellenen': guncellenen, 'hatalar': hatalar, 'mesaj': f'{eklenen} yeni, {guncellenen} güncellendi.'})


@app.route('/api/yedekle', methods=['GET'])
@token_required
def yedekle(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for table in ['hareketler', 'stoklar', 'musteriler', 'kullanicilar', 'palet_tipleri', 'makbuzlar', 'makbuz_detaylari']:
            cursor.execute(f"SELECT * FROM {table}")
            rows = cursor.fetchall()
            if rows:
                csv_buffer = BytesIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow([desc[0] for desc in cursor.description])
                csv_writer.writerows(rows)
                zipf.writestr(f'{table}.csv', csv_buffer.getvalue().decode('utf-8'))
        info = f"Palet Takip Sistemi Yedeği\nOluşturma: {get_now().strftime('%d.%m.%Y %H:%M:%S')}\nOluşturan: {current_user['ad_soyad']}"
        zipf.writestr('yedek_bilgi.txt', info)
    cursor.close()
    conn.close()
    buffer.seek(0)
    return send_file(buffer, mimetype='application/zip', as_attachment=True, download_name=f'palet_takip_yedek_{get_now().strftime("%Y%m%d_%H%M%S")}.zip')


@app.route('/api/yedekleme_ayarla', methods=['POST'])
@token_required
def yedekleme_ayarla(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json()
    aktif, periyot, saat = data.get('aktif', False), data.get('periyot', 'gunluk'), data.get('saat', '03:00')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO ayarlar (key, value) VALUES (%s, %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", ('yedekleme_aktif', str(aktif)))
    cursor.execute("INSERT INTO ayarlar (key, value) VALUES (%s, %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", ('yedekleme_periyot', periyot))
    cursor.execute("INSERT INTO ayarlar (key, value) VALUES (%s, %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", ('yedekleme_saat', saat))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({'success': True, 'mesaj': f'Otomatik yedekleme {"açıldı" if aktif else "kapatıldı"}. Periyot: {periyot}, Saat: {saat}'})


@app.route('/api/yedekleme_ayarlari', methods=['GET'])
@token_required
def yedekleme_ayarlari(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT key, value FROM ayarlar WHERE key LIKE 'yedekleme_%'")
    sonuc = cursor.fetchall()
    cursor.close()
    conn.close()
    ayarlar = {row[0].replace('yedekleme_', ''): row[1] for row in sonuc}
    return jsonify({'aktif': ayarlar.get('aktif', 'False') == 'True', 'periyot': ayarlar.get('periyot', 'gunluk'), 'saat': ayarlar.get('saat', '03:00')})


@app.route('/api/hareketleri_sifirla', methods=['POST'])
@token_required
def hareketleri_sifirla(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM makbuz_detaylari")
        cursor.execute("DELETE FROM makbuzlar")
        cursor.execute("DELETE FROM hareketler")
        cursor.execute("UPDATE stoklar SET miktar = 0")
        conn.commit()
        return jsonify({'success': True, 'mesaj': 'Tüm hareketler, makbuzlar ve stoklar başarıyla sıfırlandı!'})
    except Exception as e:
        conn.rollback()
        return jsonify({'hata': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/api/stok_excel_yukle', methods=['POST'])
@token_required
def stok_excel_yukle(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    if 'file' not in request.files:
        return jsonify({'hata': 'Dosya bulunamadı'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'hata': 'Dosya seçilmedi'}), 400
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
    except Exception as e:
        return jsonify({'hata': f'Excel okunamadı: {str(e)}'}), 400
    def std_text(text):
        if text is None: return ""
        return str(text).strip().upper().replace('İ', 'I').replace('Ş', 'S').replace('Ğ', 'G').replace('Ü', 'U').replace('Ö', 'O').replace('Ç', 'C')
    headers = [std_text(sheet.cell(row=1, column=col).value) for col in range(1, sheet.max_column + 1)]
    required_columns = ['STOK SAHIBI', 'STOK KODU', 'MIKTAR']
    for col in required_columns:
        if col not in headers:
            return jsonify({'hata': f"'{col}' sütunu bulunamadı. Lütfen dışa aktarılan 'Stoklar' rapor formatını bozmadığınızdan emin olun."}), 400
    col_indices = {h: i+1 for i, h in enumerate(headers)}
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT stok_kodu, id FROM palet_tipleri')
    paletler = {std_text(p[0]): p[1] for p in cursor.fetchall()}
    cursor.execute("SELECT ad_soyad, id FROM kullanicilar WHERE tip = 'DAGITICI'")
    dagiticilar = {std_text(d[0]): d[1] for d in cursor.fetchall()}
    cursor.execute('SELECT musteri_kodu, id FROM musteriler')
    musteriler = {std_text(m[0]): m[1] for m in cursor.fetchall()}
    guncellenen = 0
    hatalar = []
    for row in range(2, sheet.max_row + 1):
        sahip_str_raw = str(sheet.cell(row=row, column=col_indices['STOK SAHIBI']).value or '').strip()
        sahip_str = std_text(sahip_str_raw)
        palet_kodu_raw = str(sheet.cell(row=row, column=col_indices['STOK KODU']).value or '').strip()
        palet_kodu = std_text(palet_kodu_raw)
        miktar_str = sheet.cell(row=row, column=col_indices['MIKTAR']).value
        if not sahip_str or not palet_kodu or miktar_str is None:
            continue
        try:
            miktar = int(miktar_str)
        except ValueError:
            hatalar.append(f"Satır {row}: Miktar sayı formatında değil")
            continue
        if palet_kodu not in paletler:
            hatalar.append(f"Satır {row}: Geçersiz palet kodu ({palet_kodu_raw})")
            continue
        palet_tipi_id = paletler[palet_kodu]
        if sahip_str == 'DEPO':
            sahip_tipi_db = 'DEPO'
            stok_sahibi_id = 0
        elif " - " in sahip_str_raw: 
            musteri_kodu_raw = sahip_str_raw.split(" - ")[0].strip()
            musteri_kodu = std_text(musteri_kodu_raw)
            if musteri_kodu not in musteriler:
                hatalar.append(f"Satır {row}: Müşteri bulunamadı ({musteri_kodu_raw})")
                continue
            sahip_tipi_db = 'MUSTERI'
            stok_sahibi_id = musteriler[musteri_kodu]
        else: 
            if sahip_str not in dagiticilar:
                hatalar.append(f"Satır {row}: Dağıtıcı bulunamadı ({sahip_str_raw})")
                continue
            sahip_tipi_db = 'DAGITICI'
            stok_sahibi_id = dagiticilar[sahip_str]
        try:
            cursor.execute("SELECT id FROM stoklar WHERE stok_sahibi_tip = %s AND stok_sahibi_id = %s AND palet_tipi_id = %s", (sahip_tipi_db, stok_sahibi_id, palet_tipi_id))
            mevcut = cursor.fetchone()
            if mevcut:
                cursor.execute("UPDATE stoklar SET miktar = %s WHERE id = %s", (miktar, mevcut[0]))
            else:
                cursor.execute("INSERT INTO stoklar (stok_sahibi_tip, stok_sahibi_id, palet_tipi_id, miktar) VALUES (%s, %s, %s, %s)", (sahip_tipi_db, stok_sahibi_id, palet_tipi_id, miktar))
            guncellenen += 1
        except Exception as e:
            hatalar.append(f"Satır {row}: Veritabanı hatası - {str(e)}")
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({'success': True, 'guncellenen': guncellenen, 'hatalar': hatalar, 'mesaj': f'{guncellenen} stok kaydı başarıyla eşitlendi.'})


@app.route('/api/rapor/dashboard', methods=['POST'])
@token_required
def rapor_dashboard(current_user):
    if current_user['tip'] not in ('DEPOCU', 'FORKLIFT'):
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json() or {}
    baslangic = data.get('baslangic_tarihi')
    bitis = data.get('bitis_tarihi')
    if not baslangic or not bitis:
        bugun = get_now()
        baslangic = bugun.replace(day=1).strftime('%Y-%m-%d')
        bitis = bugun.strftime('%Y-%m-%d')
    baslangic_tam = f"{baslangic} 00:00:00"
    bitis_tam = f"{bitis} 23:59:59"
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT SUM(miktar) FROM stoklar WHERE stok_sahibi_tip = 'DEPO'")
        depo_stok = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT SUM(miktar) FROM stoklar WHERE stok_sahibi_tip = 'DAGITICI'")
        dagitici_stok = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT SUM(miktar) FROM stoklar WHERE stok_sahibi_tip = 'MUSTERI'")
        musteri_stok = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT SUM(miktar) FROM hareketler WHERE hareket_tipi = 'DAGITICI_MUSTERI' AND tarih >= %s AND tarih <= %s", (baslangic_tam, bitis_tam))
        verilen = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT SUM(miktar) FROM hareketler WHERE hareket_tipi = 'MUSTERI_DAGITICI' AND tarih >= %s AND tarih <= %s", (baslangic_tam, bitis_tam))
        toplanan = int(cursor.fetchone()[0] or 0)
        cursor.execute('''
            SELECT m.musteri_adi, SUM(s.miktar) as toplam
            FROM stoklar s JOIN musteriler m ON s.stok_sahibi_id = m.id
            WHERE s.stok_sahibi_tip = 'MUSTERI' AND s.miktar > 0
            GROUP BY m.id, m.musteri_adi ORDER BY toplam DESC LIMIT 10
        ''')
        bekleyenler = [{'ad': r[0], 'miktar': int(r[1] or 0)} for r in cursor.fetchall()]
        cursor.execute('''
            SELECT u.ad_soyad, 
                   SUM(CASE WHEN h.hareket_tipi = 'DAGITICI_MUSTERI' THEN h.miktar ELSE 0 END) as v,
                   SUM(CASE WHEN h.hareket_tipi = 'MUSTERI_DAGITICI' THEN h.miktar ELSE 0 END) as t
            FROM hareketler h JOIN kullanicilar u ON h.yapan_kullanici_id = u.id
            WHERE u.tip = 'DAGITICI' AND h.tarih >= %s AND h.tarih <= %s
            GROUP BY u.id, u.ad_soyad ORDER BY t DESC
        ''', (baslangic_tam, bitis_tam))
        perf = [{'ad': r[0], 'verilen': int(r[1] or 0), 'toplanan': int(r[2] or 0)} for r in cursor.fetchall()]
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'hata': f'Sistemsel Veri Hatası: {str(e)}'}), 500
    cursor.close()
    conn.close()
    oran = round((toplanan / verilen * 100), 1) if verilen > 0 else 0
    return jsonify({
        'stok': {'depo': depo_stok, 'dagitici': dagitici_stok, 'musteri': musteri_stok, 'toplam': depo_stok+dagitici_stok+musteri_stok},
        'verim': {'v': verilen, 't': toplanan, 'oran': oran},
        'bekleyen': bekleyenler,
        'perf': perf
    })


@app.route('/api/toplanacak_paletler', methods=['GET'])
@token_required
def toplanacak_paletler(current_user):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            SELECT m.id, m.musteri_kodu, m.musteri_adi, SUM(s.miktar) as toplam_stok
            FROM stoklar s
            JOIN musteriler m ON s.stok_sahibi_id = m.id
            WHERE s.stok_sahibi_tip = 'MUSTERI' AND s.miktar > 0
            GROUP BY m.id, m.musteri_kodu, m.musteri_adi
        ''')
        musteri_stoklari = cursor.fetchall()
        bugun = get_now()
        sonuclar = []
        for m in musteri_stoklari:
            m_id, m_kodu, m_adi, toplam_stok = m
            toplam_stok = int(toplam_stok)
            cursor.execute('''
                SELECT tarih, miktar, yapan_kullanici_id
                FROM hareketler
                WHERE hareket_tipi = 'DAGITICI_MUSTERI' AND alan_id = %s
                ORDER BY tarih DESC
            ''', (m_id,))
            verme_hareketleri = cursor.fetchall()
            yas_0_7 = 0
            yas_8_14 = 0
            yas_15_21 = 0
            yas_22_arti = 0
            kalan_stok = toplam_stok
            ilgili_dagitici_mi = False
            for h in verme_hareketleri:
                if kalan_stok <= 0:
                    break
                h_tarih_raw = h[0]
                h_miktar = int(h[1])
                h_dagitici_id = h[2]
                if current_user['tip'] == 'DAGITICI' and h_dagitici_id == current_user['id']:
                    ilgili_dagitici_mi = True
                if isinstance(h_tarih_raw, datetime):
                    h_tarih = h_tarih_raw
                else:
                    try:
                        h_tarih = datetime.strptime(str(h_tarih_raw).split('.')[0], '%Y-%m-%d %H:%M:%S')
                    except:
                        h_tarih = bugun
                fark_gun = (bugun - h_tarih).days
                islem_miktari = min(kalan_stok, h_miktar)
                kalan_stok -= islem_miktari
                if fark_gun <= 7:
                    yas_0_7 += islem_miktari
                elif fark_gun <= 14:
                    yas_8_14 += islem_miktari
                elif fark_gun <= 21:
                    yas_15_21 += islem_miktari
                else:
                    yas_22_arti += islem_miktari
            if kalan_stok > 0:
                yas_22_arti += kalan_stok
            # Dağıtıcı bazlı kırılım hesapla
            dagitici_kirilimleri = {}
            kalan_stok2 = toplam_stok
            for h in verme_hareketleri:
                if kalan_stok2 <= 0:
                    break
                h_tarih_raw = h[0]
                h_miktar = int(h[1])
                h_dagitici_id = h[2]
                if isinstance(h_tarih_raw, datetime):
                    h_tarih2 = h_tarih_raw
                else:
                    try:
                        h_tarih2 = datetime.strptime(str(h_tarih_raw).split('.')[0], '%Y-%m-%d %H:%M:%S')
                    except:
                        h_tarih2 = bugun
                fark_gun2 = (bugun - h_tarih2).days
                islem_miktari2 = min(kalan_stok2, h_miktar)
                kalan_stok2 -= islem_miktari2
                if h_dagitici_id not in dagitici_kirilimleri:
                    dagitici_kirilimleri[h_dagitici_id] = {'g0_7': 0, 'g8_14': 0, 'g15_21': 0, 'g22': 0, 'toplam': 0}
                dagitici_kirilimleri[h_dagitici_id]['toplam'] += islem_miktari2
                if fark_gun2 <= 7:
                    dagitici_kirilimleri[h_dagitici_id]['g0_7'] += islem_miktari2
                elif fark_gun2 <= 14:
                    dagitici_kirilimleri[h_dagitici_id]['g8_14'] += islem_miktari2
                elif fark_gun2 <= 21:
                    dagitici_kirilimleri[h_dagitici_id]['g15_21'] += islem_miktari2
                else:
                    dagitici_kirilimleri[h_dagitici_id]['g22'] += islem_miktari2
            # Dağıtıcı adlarını getir
            dagitici_listesi = []
            for dag_id, dag_data in dagitici_kirilimleri.items():
                cursor.execute("SELECT ad_soyad FROM kullanicilar WHERE id = %s", (dag_id,))
                dag_row = cursor.fetchone()
                dag_ad = dag_row[0] if dag_row else f"ID:{dag_id}"
                dagitici_listesi.append({'dagitici_ad': dag_ad, 'dagitici_id': dag_id, **dag_data})

            if current_user['tip'] in ('DEPOCU', 'FORKLIFT') or (current_user['tip'] == 'DAGITICI' and ilgili_dagitici_mi):
                if (yas_8_14 + yas_15_21 + yas_22_arti) > 0:
                    sonuclar.append({
                        'musteri': f"{m_kodu} - {m_adi}",
                        'musteri_id': m_id,
                        'toplam': toplam_stok,
                        'g0_7': yas_0_7,
                        'g8_14': yas_8_14,
                        'g15_21': yas_15_21,
                        'g22': yas_22_arti,
                        'dagiticilar': dagitici_listesi
                    })
        sonuclar.sort(key=lambda x: (x['g22'], x['g15_21'], x['g8_14']), reverse=True)
        return jsonify(sonuclar)
    except Exception as e:
        return jsonify({'hata': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


# ============================================================
# BİLDİRİM SİSTEMİ - AYARLAR API
# ============================================================

def ayar_getir(key, varsayilan=''):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT value FROM ayarlar WHERE key = %s", (key,))
        row = cursor.fetchone()
        return row[0] if row else varsayilan
    except:
        return varsayilan
    finally:
        cursor.close()
        conn.close()

def ayar_kaydet(key, value):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO ayarlar (key, value) VALUES (%s, %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
            (key, value)
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cursor.close()
        conn.close()


@app.route('/api/bildirim_ayarlari', methods=['GET'])
@token_required
def bildirim_ayarlari_getir(current_user):
    if current_user['tip'] != 'DEPOCU':
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    ayarlar = {
        'email_aktif':      ayar_getir('bildirim_email_aktif', '0'),
        'email_gonderen':   ayar_getir('bildirim_email_gonderen', ''),
        'email_sifre':      ayar_getir('bildirim_email_sifre', ''),
        'email_smtp':       ayar_getir('bildirim_email_smtp', 'smtp.gmail.com'),
        'email_port':       ayar_getir('bildirim_email_port', '587'),
        'email_alicilar':   ayar_getir('bildirim_email_alicilar', ''),
        'whatsapp_aktif':   ayar_getir('bildirim_whatsapp_aktif', '0'),
        'twilio_sid':       ayar_getir('bildirim_twilio_sid', ''),
        'twilio_token':     ayar_getir('bildirim_twilio_token', ''),
        'twilio_from':      ayar_getir('bildirim_twilio_from', ''),
        'whatsapp_alicilar':ayar_getir('bildirim_whatsapp_alicilar', ''),
        'bildirim_saati':   ayar_getir('bildirim_saati', '08:00'),
        'min_gun':          ayar_getir('bildirim_min_gun', '21'),
    }
    return jsonify(ayarlar)


@app.route('/api/bildirim_ayarlari', methods=['POST'])
@token_required
def bildirim_ayarlari_kaydet(current_user):
    if current_user['tip'] != 'DEPOCU':
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    data = request.get_json() or {}
    mapping = {
        'email_aktif':       'bildirim_email_aktif',
        'email_gonderen':    'bildirim_email_gonderen',
        'email_sifre':       'bildirim_email_sifre',
        'email_smtp':        'bildirim_email_smtp',
        'email_port':        'bildirim_email_port',
        'email_alicilar':    'bildirim_email_alicilar',
        'whatsapp_aktif':    'bildirim_whatsapp_aktif',
        'twilio_sid':        'bildirim_twilio_sid',
        'twilio_token':      'bildirim_twilio_token',
        'twilio_from':       'bildirim_twilio_from',
        'whatsapp_alicilar': 'bildirim_whatsapp_alicilar',
        'bildirim_saati':    'bildirim_saati',
        'min_gun':           'bildirim_min_gun',
    }
    for alan, key in mapping.items():
        if alan in data:
            ayar_kaydet(key, str(data[alan]))
    return jsonify({'success': True, 'mesaj': 'Ayarlar kaydedildi'})


# ============================================================
# RAPOR VERİSİ HAZIRLA (22+ gün, dağıtıcı + palet tipi detaylı)
# ============================================================

def bildirim_raporu_hazirla(min_gun=21):
    """22+ gün (veya min_gun üzeri) bekleyen paletleri dağıtıcı bazlı, palet tipi detaylı döndür."""
    conn = get_db_connection()
    cursor = conn.cursor()
    bugun = get_now()
    try:
        cursor.execute('''
            SELECT m.id, m.musteri_kodu, m.musteri_adi,
                   pt.stok_kodu, pt.palet_adi, s.miktar
            FROM stoklar s
            JOIN musteriler m ON s.stok_sahibi_id = m.id
            JOIN palet_tipleri pt ON s.palet_tipi_id = pt.id
            WHERE s.stok_sahibi_tip = 'MUSTERI' AND s.miktar > 0
            ORDER BY m.id, pt.id
        ''')
        musteri_stoklar = cursor.fetchall()

        # {musteri_id: {palet_tipi_id: miktar}}
        musteri_palet_map = {}
        for row in musteri_stoklar:
            m_id, m_kodu, m_adi, pt_kod, pt_ad, miktar = row
            if m_id not in musteri_palet_map:
                musteri_palet_map[m_id] = {
                    'musteri_kodu': m_kodu,
                    'musteri_adi': m_adi,
                    'palet_stoklar': {}
                }
            musteri_palet_map[m_id]['palet_stoklar'][pt_kod] = {
                'ad': pt_ad, 'miktar': int(miktar)
            }

        # Her müşteri için hareketleri getir, gün hesapla, dağıtıcı kır
        dagitici_rapor = {}  # {dagitici_id: {ad, musteriler: [...]}}

        for m_id, m_info in musteri_palet_map.items():
            toplam_stok = sum(v['miktar'] for v in m_info['palet_stoklar'].values())
            if toplam_stok <= 0:
                continue

            cursor.execute('''
                SELECT tarih, miktar, yapan_kullanici_id, palet_tipi_id
                FROM hareketler
                WHERE hareket_tipi = 'DAGITICI_MUSTERI' AND alan_id = %s
                ORDER BY tarih DESC
            ''', (m_id,))
            hareketler = cursor.fetchall()

            kalan_stok = toplam_stok
            # {dagitici_id: {palet_tipi_id: {gun_grubu: miktar}}}
            dag_palet_gun = {}

            for h in hareketler:
                if kalan_stok <= 0:
                    break
                h_tarih_raw, h_miktar, h_dag_id, h_palet_id = h[0], int(h[1]), h[2], h[3]
                if isinstance(h_tarih_raw, datetime):
                    h_tarih = h_tarih_raw
                else:
                    try:
                        h_tarih = datetime.strptime(str(h_tarih_raw).split('.')[0], '%Y-%m-%d %H:%M:%S')
                    except:
                        h_tarih = bugun
                fark_gun = (bugun - h_tarih).days
                islem = min(kalan_stok, h_miktar)
                kalan_stok -= islem

                if fark_gun <= min_gun:
                    continue  # Sadece min_gun üzerini al

                if h_dag_id not in dag_palet_gun:
                    dag_palet_gun[h_dag_id] = {}
                key = str(h_palet_id)
                if key not in dag_palet_gun[h_dag_id]:
                    dag_palet_gun[h_dag_id][key] = 0
                dag_palet_gun[h_dag_id][key] += islem

            # Dağıtıcıların adlarını al ve raporu oluştur
            for dag_id, palet_data in dag_palet_gun.items():
                if not palet_data:
                    continue
                if dag_id not in dagitici_rapor:
                    cursor.execute("SELECT ad_soyad FROM kullanicilar WHERE id = %s", (dag_id,))
                    dag_row = cursor.fetchone()
                    dagitici_rapor[dag_id] = {
                        'dagitici_ad': dag_row[0] if dag_row else f'ID:{dag_id}',
                        'musteriler': []
                    }

                # Palet tipi adlarını getir
                palet_detay = []
                toplam_musteri = 0
                for pt_id_str, adet in palet_data.items():
                    if adet <= 0:
                        continue
                    cursor.execute("SELECT stok_kodu, palet_adi FROM palet_tipleri WHERE id = %s", (int(pt_id_str),))
                    pt_row = cursor.fetchone()
                    pt_ad = f"{pt_row[0]} - {pt_row[1]}" if pt_row else f"Tip:{pt_id_str}"
                    palet_detay.append({'palet': pt_ad, 'adet': adet})
                    toplam_musteri += adet

                if toplam_musteri > 0:
                    dagitici_rapor[dag_id]['musteriler'].append({
                        'musteri_kodu': m_info['musteri_kodu'],
                        'musteri_adi': m_info['musteri_adi'],
                        'toplam': toplam_musteri,
                        'paletler': palet_detay
                    })

        # Sadece müşterisi olan dağıtıcıları döndür
        return {dag_id: v for dag_id, v in dagitici_rapor.items() if v['musteriler']}
    finally:
        cursor.close()
        conn.close()


def rapor_html_olustur(rapor, min_gun=21):
    tarih_str = get_now().strftime('%d.%m.%Y %H:%M')
    toplam_musteri = sum(len(v['musteriler']) for v in rapor.values())
    toplam_palet = sum(
        sum(m['toplam'] for m in v['musteriler']) for v in rapor.values()
    )

    dag_html = ''
    for dag_data in sorted(rapor.values(), key=lambda x: x['dagitici_ad']):
        dag_toplam = sum(m['toplam'] for m in dag_data['musteriler'])
        musteri_rows = ''
        for m in sorted(dag_data['musteriler'], key=lambda x: -x['toplam']):
            palet_str = ', '.join(f"<b>{p['adet']} adet</b> {p['palet']}" for p in m['paletler'])
            musteri_rows += f'''
            <tr>
              <td style="padding:8px 12px; border-bottom:1px solid #eee;">{m["musteri_kodu"]} - {m["musteri_adi"]}</td>
              <td style="padding:8px 12px; border-bottom:1px solid #eee; text-align:center; font-weight:700; color:#b71c1c;">{m["toplam"]}</td>
              <td style="padding:8px 12px; border-bottom:1px solid #eee; font-size:0.9em; color:#555;">{palet_str}</td>
            </tr>'''

        dag_html += f'''
        <div style="margin-bottom:24px; border:1px solid #ddd; border-radius:8px; overflow:hidden;">
          <div style="background:#1565c0; color:white; padding:10px 16px; display:flex; justify-content:space-between; align-items:center;">
            <span style="font-size:1rem; font-weight:700;">🚚 {dag_data["dagitici_ad"]}</span>
            <span style="background:rgba(255,255,255,0.2); padding:3px 10px; border-radius:12px; font-size:0.85rem;">Toplam: {dag_toplam} palet</span>
          </div>
          <table style="width:100%; border-collapse:collapse; font-size:0.9rem;">
            <thead>
              <tr style="background:#e3f2fd;">
                <th style="padding:8px 12px; text-align:left; color:#1565c0;">Müşteri</th>
                <th style="padding:8px 12px; text-align:center; color:#1565c0; width:80px;">Adet</th>
                <th style="padding:8px 12px; text-align:left; color:#1565c0;">Palet Tipi</th>
              </tr>
            </thead>
            <tbody>{musteri_rows}</tbody>
          </table>
        </div>'''

    return f'''<!DOCTYPE html>
<html lang="tr">
<head><meta charset="UTF-8">
<title>Palet Toplama Raporu - {tarih_str}</title>
</head>
<body style="font-family:Arial,sans-serif; background:#f5f5f5; margin:0; padding:20px;">
<div style="max-width:700px; margin:0 auto; background:white; border-radius:12px; overflow:hidden; box-shadow:0 2px 12px rgba(0,0,0,0.1);">
  <div style="background:#b71c1c; color:white; padding:20px 24px;">
    <h2 style="margin:0 0 6px;">🚨 Palet Toplama Raporu</h2>
    <p style="margin:0; opacity:0.9; font-size:0.9rem;">{tarih_str} tarihli günlük rapor — {min_gun}+ gün bekleyen paletler</p>
  </div>
  <div style="padding:16px 24px; background:#fff3e0; border-bottom:1px solid #ffe0b2; display:flex; gap:24px;">
    <div><span style="font-size:0.75rem; color:#e65100;">TOPLAM MÜŞTERİ</span><br><strong style="font-size:1.4rem; color:#b71c1c;">{toplam_musteri}</strong></div>
    <div><span style="font-size:0.75rem; color:#e65100;">TOPLANACAK PALET</span><br><strong style="font-size:1.4rem; color:#b71c1c;">{toplam_palet}</strong></div>
    <div><span style="font-size:0.75rem; color:#e65100;">DAĞITICI SAYISI</span><br><strong style="font-size:1.4rem; color:#b71c1c;">{len(rapor)}</strong></div>
  </div>
  <div style="padding:20px 24px;">
    {dag_html if dag_html else '<p style="color:#4caf50; text-align:center; padding:20px;">✅ Toplanması gereken palet bulunmuyor.</p>'}
  </div>
  <div style="padding:12px 24px; background:#f5f5f5; text-align:center; font-size:0.75rem; color:#999;">
    Bu rapor Palet Takip Sistemi tarafından otomatik oluşturulmuştur.
  </div>
</div>
</body></html>'''


def rapor_text_olustur(rapor, min_gun=21):
    tarih_str = get_now().strftime('%d.%m.%Y %H:%M')
    lines = [f"🚨 PALET TOPLAMA RAPORU - {tarih_str}",
             f"({min_gun}+ gün bekleyen paletler)", "=" * 45]
    for dag_data in sorted(rapor.values(), key=lambda x: x['dagitici_ad']):
        dag_toplam = sum(m['toplam'] for m in dag_data['musteriler'])
        lines.append(f"\n🚚 {dag_data['dagitici_ad']} — {dag_toplam} palet")
        lines.append("-" * 35)
        for m in sorted(dag_data['musteriler'], key=lambda x: -x['toplam']):
            lines.append(f"  📍 {m['musteri_kodu']} - {m['musteri_adi']}")
            for p in m['paletler']:
                lines.append(f"     • {p['adet']} adet {p['palet']}")
    lines.append("\n" + "=" * 45)
    lines.append("Palet Takip Sistemi - Otomatik Rapor")
    return "\n".join(lines)


# ============================================================
# E-POSTA GÖNDER
# ============================================================

def email_gonder(rapor, min_gun=21):
    aktif = ayar_getir('bildirim_email_aktif', '0')
    if aktif != '1':
        return False, 'E-posta bildirimi aktif değil'
    gonderen    = ayar_getir('bildirim_email_gonderen', '')
    sifre       = ayar_getir('bildirim_email_sifre', '')
    smtp_host   = ayar_getir('bildirim_email_smtp', 'smtp.gmail.com')
    smtp_port   = int(ayar_getir('bildirim_email_port', '587'))
    alicilar_str= ayar_getir('bildirim_email_alicilar', '')
    if not gonderen or not sifre or not alicilar_str:
        return False, 'E-posta ayarları eksik'
    alicilar = [a.strip() for a in alicilar_str.split(',') if a.strip()]
    tarih_str = get_now().strftime('%d.%m.%Y')
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"🚨 Palet Toplama Raporu - {tarih_str}"
        msg['From'] = gonderen
        msg['To'] = ', '.join(alicilar)
        html_icerik = rapor_html_olustur(rapor, min_gun)
        text_icerik = rapor_text_olustur(rapor, min_gun)
        msg.attach(MIMEText(text_icerik, 'plain', 'utf-8'))
        msg.attach(MIMEText(html_icerik, 'html', 'utf-8'))
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
            server.ehlo()
            server.starttls()
            server.login(gonderen, sifre)
            server.sendmail(gonderen, alicilar, msg.as_string())
        return True, f"{len(alicilar)} alıcıya e-posta gönderildi"
    except Exception as e:
        return False, f"E-posta hatası: {str(e)}"


# ============================================================
# WHATSAPP GÖNDER (Twilio)
# ============================================================

def whatsapp_gonder(rapor, min_gun=21):
    aktif = ayar_getir('bildirim_whatsapp_aktif', '0')
    if aktif != '1':
        return False, 'WhatsApp bildirimi aktif değil'
    try:
        from twilio.rest import Client
    except ImportError:
        return False, 'Twilio kütüphanesi yüklü değil (pip install twilio)'
    sid        = ayar_getir('bildirim_twilio_sid', '')
    token      = ayar_getir('bildirim_twilio_token', '')
    from_num   = ayar_getir('bildirim_twilio_from', '')
    alicilar_str = ayar_getir('bildirim_whatsapp_alicilar', '')
    if not sid or not token or not from_num or not alicilar_str:
        return False, 'WhatsApp (Twilio) ayarları eksik'
    alicilar = [a.strip() for a in alicilar_str.split(',') if a.strip()]
    mesaj = rapor_text_olustur(rapor, min_gun)
    if len(mesaj) > 1500:
        mesaj = mesaj[:1497] + '...'
    client = Client(sid, token)
    basarili, hatali = 0, 0
    for alici in alicilar:
        try:
            client.messages.create(
                from_=f"whatsapp:{from_num}",
                to=f"whatsapp:{alici}",
                body=mesaj
            )
            basarili += 1
        except Exception:
            hatali += 1
    return True, f"WhatsApp: {basarili} gönderildi, {hatali} başarısız"


# ============================================================
# BİLDİRİM GÖNDER (her iki kanal)
# ============================================================

def bildirim_gonder_tum():
    min_gun = int(ayar_getir('bildirim_min_gun', '21'))
    rapor = bildirim_raporu_hazirla(min_gun)
    sonuclar = []
    email_ok, email_msg = email_gonder(rapor, min_gun)
    sonuclar.append({'kanal': 'E-posta', 'basarili': email_ok, 'mesaj': email_msg})
    wp_ok, wp_msg = whatsapp_gonder(rapor, min_gun)
    sonuclar.append({'kanal': 'WhatsApp', 'basarili': wp_ok, 'mesaj': wp_msg})
    return sonuclar, len(rapor)


@app.route('/api/bildirim_gonder', methods=['POST'])
@token_required
def bildirim_gonder_endpoint(current_user):
    if current_user['tip'] != 'DEPOCU':
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    try:
        sonuclar, dag_sayisi = bildirim_gonder_tum()
        return jsonify({'success': True, 'dagitici_sayisi': dag_sayisi, 'sonuclar': sonuclar})
    except Exception as e:
        return jsonify({'hata': str(e)}), 500


@app.route('/api/bildirim_onizleme', methods=['GET'])
@token_required
def bildirim_onizleme(current_user):
    if current_user['tip'] != 'DEPOCU':
        return jsonify({'hata': 'Yetkisiz erişim'}), 403
    try:
        min_gun = int(ayar_getir('bildirim_min_gun', '21'))
        rapor = bildirim_raporu_hazirla(min_gun)
        ozet = []
        for dag_data in sorted(rapor.values(), key=lambda x: x['dagitici_ad']):
            ozet.append({
                'dagitici': dag_data['dagitici_ad'],
                'musteri_sayisi': len(dag_data['musteriler']),
                'toplam_palet': sum(m['toplam'] for m in dag_data['musteriler']),
                'musteriler': dag_data['musteriler']
            })
        return jsonify({'ozet': ozet, 'dagitici_sayisi': len(ozet), 'min_gun': min_gun})
    except Exception as e:
        return jsonify({'hata': str(e)}), 500


# ============================================================
# OTOMATİK ZAMANLAYICI (Thread tabanlı)
# ============================================================

_zamanlayici_thread = None
_zamanlayici_aktif = False

def zamanlayici_dongu():
    global _zamanlayici_aktif
    son_gonderim_tarihi = None
    while _zamanlayici_aktif:
        try:
            bildirim_saati = ayar_getir('bildirim_saati', '08:00')
            simdi = get_now()
            bugun_str = simdi.strftime('%Y-%m-%d')
            saat_str = simdi.strftime('%H:%M')
            if saat_str == bildirim_saati and son_gonderim_tarihi != bugun_str:
                email_aktif = ayar_getir('bildirim_email_aktif', '0')
                wp_aktif = ayar_getir('bildirim_whatsapp_aktif', '0')
                if email_aktif == '1' or wp_aktif == '1':
                    bildirim_gonder_tum()
                    son_gonderim_tarihi = bugun_str
        except Exception:
            pass
        time.sleep(55)  # Her 55 saniyede kontrol et

def zamanlayici_baslat():
    global _zamanlayici_thread, _zamanlayici_aktif
    if _zamanlayici_thread and _zamanlayici_thread.is_alive():
        return
    _zamanlayici_aktif = True
    _zamanlayici_thread = threading.Thread(target=zamanlayici_dongu, daemon=True)
    _zamanlayici_thread.start()



if __name__ == '__main__':
    veritabani_olustur()
    zamanlayici_baslat()
    from waitress import serve
    port = int(os.environ.get('PORT', 5000))
    serve(app, host='0.0.0.0', port=port, threads=4)
